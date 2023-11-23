import openpyxl

# Carregue o arquivo Excel
workbook = openpyxl.load_workbook('Ras.xlsx')
sheet = workbook.active  # Ou selecione a planilha desejada

# Coluna na qual você deseja verificar a validação
coluna_alvo = 'A'  # Substitua pela letra da coluna desejada

# Verifique as células na coluna alvo
linhas_com_informacao = []

informacao_procurada = input("Digite a informação que você está procurando: ")

# Calcule o número da coluna correspondente
coluna_numero = openpyxl.utils.column_index_from_string(coluna_alvo)

# Itere pelas linhas a partir da segunda linha (min_row=2)
for row in sheet.iter_rows(min_row=2, min_col=coluna_numero, max_col=coluna_numero):
    cell = row[0]  # Acesse a primeira (e única) célula na tupla
    if cell.value and informacao_procurada.upper() in str(cell.value).upper():
        linhas_com_informacao.append(cell.row)


# Verifique se há linhas encontradas
if linhas_com_informacao:
    linha_validada = linhas_com_informacao[0]
    print("Linha com a informação procurada:", linha_validada)
    
    # Acesse a informação da coluna B na linha validada
    informacao_coluna_B = sheet.cell(row=linha_validada, column=2).value
    
    print("Valor da coluna B na linha", linha_validada, ":", informacao_coluna_B)
else:
    print("Nenhuma linha com a informação procurada foi encontrada")

# Feche o arquivo do Excel
workbook.close()
